import selenium
from selenium.webdriver.common.by import By
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from openpyxl import Workbook

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

import time
import datetime
import requests


import os
import pandas as pd
from tqdm import tqdm 


"""
네이버 리뷰에서 음식점을 정하고
해당 음식점에 리류를 작성한 유저의 닉네임과 링크를 크롤링
이후 엑셀에 저장

"""
def function1(url):
    cnt = 1

    #엑셀 생성
    xlsx = Workbook()
    
    #output이라는 sheet에 저장해준다.
    list_sheet = xlsx.create_sheet('output')
    list_sheet.append(['nickname', 'link'])
    now = datetime.datetime.now()
    
    #크롤링중에 창 안뜨고 진행할 수 있도록 한다
    options = webdriver.ChromeOptions()
    options.add_argument("headless")

    
    try:
        driver = webdriver.Chrome(options=options)
        #음식점 링크를 받아온다
        driver.get(url)
        driver.implicitly_wait(2)

        count = 0
        try:
            # 해당 음식점에 리뷰를 작성한 모든 리뷰어의 개인 프로필 링크를 크롤링 한다.
            while True:
                count+=1
                
                #XPATH를 이용해 프로필을 클릭한다.
                driver.find_element(By.XPATH, '//*[@id="app-root"]/div/div/div/div[6]/div[2]/div[3]/div[2]/div/a/span').click()
                time.sleep(0.04)
                if count == 500:
                    break

        except Exception as e:
            print(e)
            print(count)

        time.sleep(1)
        html = driver.page_source
        bs = BeautifulSoup(html, 'lxml')
        reviews = bs.select('li.YeINN')

        #해당 리뷰어의 리뷰 목록에 접근하는 과정
        for r in reviews:
            nickname = r.select_one('div.VYGLG')
            link=r.select_one('a.p24Ki')['href']

            # exception handling
            nickname = nickname.text if nickname else ''    
            
            #엑셀 파일에 리뷰어 nickname과 링크를 저장한다.
            list_sheet.append([nickname, link])
            time.sleep(0.06)    

        driver.quit()
        time.sleep(0.06)
        #save file 
        
        file_name = 'reviewer_link_' + now.strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx'
        xlsx.save(file_name)
        
        cnt+=1

    except Exception as e:
        print("function1 error occurs")
        print(e)
        
        file_name = 'reviewer_link_' + now.strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx'
        xlsx.save(file_name)


"""
리뷰어의 개인 프로필 링크 접근
작성한 모든 리뷰를 볼 수 있는 페이지 이동 

"""
def review_2(url_path):
    #리뷰어 개인 프로필의 링크
    url = url_path
    options = webdriver.ChromeOptions()
    options.add_argument("headless")


    try:
        driver = webdriver.Chrome(options = options)
        driver.get(url)

        #버튼 클릭
        reviewer_url_xpath = WebDriverWait(driver, 3).until(
    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/div[1]/div[3]/div/ul/li[1]/button")))
        reviewer_url_xpath.click()
        time.sleep(1.2)


        get_url = driver.current_url 
        driver.quit()
        return get_url
    except Exception as e:
        #print(e)
        return 0
    
"""
리뷰를 바탕으로 데이터 구축

"""
def function3(df_):
    now = datetime.datetime.now()

    #엑셀에 저장해준다
    xlsx = Workbook()
    
    #output이라는 sheet에 저장
    list_sheet = xlsx.create_sheet('output')
    
    #크롤링 할 내용
    list_sheet.append(['name','restaurant', 'content', 'type', 'address','date','url'])

    #크롤링 옵션
    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    options.add_argument('window-size=1920x1080')
    options.add_argument("disable-gpu")

    # 리뷰어 개인 프로필의 링크를 받아서 시작한다.
    count=0
    for url_ in tqdm(df['link']):
        count+=1
        
        #중간에 저장해준다
        if count % 100 ==0:
            file_name = f'naver_review_{count}_.xlsx'
            xlsx.save(file_name)
            
        try:
            url = review_2(url_)
            if url == 0:
                continue
            driver = webdriver.Chrome(options= options)
            driver.get(url)

            time.sleep(1.5)
            html = driver.page_source
            bs = BeautifulSoup(html, 'lxml')
            
            #원하는 데이터를 크롤링하기 위해 html파일에 간접적으로 접근하는 방식을 선택했다.
            name=bs.select_one('header._2nqODz>div.jA_lkM>button.wTaI4v>h1._2LIPHf')
            name=name.text if name else ''

            #리뷰어가 많은 리뷰를 작성했을 경우를 위해 스크롤해서 페이지 밑으로 내려간다.
            for i in range(50):
                driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
                time.sleep(0.04)

            html = driver.page_source
            bs = BeautifulSoup(html, 'lxml')

            #각 리뷰의 html을 잡아준다.
            reviews = bs.select('div._27tH92')

            # 리뷰에서 필요한 내용을 크롤링한다.
            for r in reviews:
                restaurant = r.select_one('div.MF77ib>button.wTaI4v>span._1QGRWW')
                content = r.select_one('div._3-ITu7')
                type = r.select('div._2vBfgu>span.wzFIfJ')[0]
                address = r.select('div._2vBfgu>span.wzFIfJ')[1]
                date=r.select('div.pOj49R>div._15xwjO>div.hol3Ic>div>span._3nNYBi>time')[0]

                # exception handling
                restaurant = restaurant.text if restaurant else ''
                content = content.text if content else ''
                type = type.text if type else ''
                address = address.text if address else ''
                date=date.text if date else ''

                list_sheet.append([name,restaurant, content, type, address, date,url])
                time.sleep(0.06)

        except:
            continue

    file_name = f'naver_review_full.xlsx'
    xlsx.save(file_name)
    driver.quit()


def read_all_excel_files_in_current_path(sheet_name='output'):
    current_path = os.getcwd()

    all_files = os.listdir(current_path)

    excel_files = [file for file in all_files if file.endswith(('.xlsx', '.xls'))]

    if not excel_files:
        print("No Excel files found in the current directory.")
        return None


    all_dataframes = []
    for excel_file in excel_files:
        file_path = os.path.join(current_path, excel_file)
        try:
            df = pd.read_excel(file_path, sheet_name)
            all_dataframes.append(df)
        except Exception as e:
            print(f"Error reading sheet '{sheet_name}' from '{excel_file}': {str(e)}")

    return all_dataframes

def concat_and_save_to_excel(dataframes, output_file='concatenated_output.xlsx'):
    
    concatenated_df = pd.concat(dataframes, ignore_index=True)


    concatenated_df.to_excel(output_file, index=False, engine='openpyxl')
    print(f"Concatenated data saved to {output_file}")



if __name__ == "__main__":
    now = datetime.datetime.now()
    print("시작시간   ",end='')
    print(now)
    current_path = os.getcwd()
    all_files = os.listdir(current_path)
    excel_files = [file for file in all_files if file.endswith(('.xlsx', '.xls'))]
    #import IPython; IPython.embed(colors='Linux'); exit(1)

    if "concatenated_output.xlsx" in all_files:
        df = pd.read_excel("")
        print("concatenated_output.xlsx already exists!")
        function3(df)
        
    elif excel_files:
        print("excel exits!.")
        result_dataframes = read_all_excel_files_in_current_path()
        concat_and_save_to_excel(result_dataframes)

        df = pd.read_excel("")
        print("concatenated_output.xlsx already exists!")
        function3(df)

    else:
        text_path = [

                     ]
                    

        #음식점 url에서 리뷰어 url 모음 - 함수1
        for url_path in text_path:
            url_path = url_path
            function1(url_path)
        print("function1 end")

        result_dataframes = read_all_excel_files_in_current_path()
        
        concat_and_save_to_excel(result_dataframes)
        df = pd.read_excel("")

        #함수3
        function3(df)


    now = datetime.datetime.now()
    print("끝난 시간   ",end='')
    print(now)